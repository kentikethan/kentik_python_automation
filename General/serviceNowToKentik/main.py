# The goal of this program is to read in a csv file with strings that should be in interface descriptions
#  and add a device Id and an interface id to the spreadsheet

import openpyxl
import datetime
from dateutil.relativedelta import relativedelta
import re

# open the interfaces workbook
interfaces_workbook = openpyxl.load_workbook('./interfaces.xlsx')

# grab the active worksheet
interfaces_ws = interfaces_workbook.active

# open the circuit information workbook
circuit_workbook = openpyxl.load_workbook('./circuits.xlsx')
# open the circuit worksheet
circuit_ws = circuit_workbook.active
# open the output information workbook
output_workbook = openpyxl.load_workbook('./connectivit_costs_bulk_import.xlsx')
# open the output worksheet
output_ws = output_workbook.active
# set the new row integer
c_number = 2
# iterate through the cells in the circuit spreadsheet and assign variables for key items
for row in circuit_ws.iter_rows(min_row=2,min_col=7,max_col=7):
    for cell in row:
        # notes cell for this row
        c_notes = circuit_ws.cell(row = cell.row, column = 15)
        # used to search the interface descriptions for a match
        circuit_id = str(circuit_ws.cell(row=cell.row, column=8).value)
        # used to pick the A side of the interface
        source_site_name = circuit_ws.cell(row=cell.row, column=1).value
        # used to filter out non point to point circuits
        connect_type = circuit_ws.cell(row=cell.row, column=3).value
        # used to get the bandwidth of the connection
        circuit_type = circuit_ws.cell(row=cell.row, column=5).value
        commit_bandwidth = circuit_ws.cell(row=cell.row, column=10).value
        match = re.search(r'(\d+\.?\d*)', str(commit_bandwidth))
        if match:
            bandwidth = match.group(1)
        else:
            bandwidth = 0
            c_notes.value = "Commit Bandwidth incorrect"
        #if circuit_type is None:
        #    c_notes.value = "circuit type was blank"
        #    continue
        #if circuit_type.find("1G") > -1:
        #    bandwidth = 1
        #elif circuit_type.find("10G") > -1:
        #    bandwidth = 10
        #elif circuit_type.find("100G") > -1:
        #    bandwidth = 100
        # used for provider selection
        provider_name = circuit_ws.cell(row=cell.row, column=6).value
        # currency used in output
        currency = circuit_ws.cell(row=cell.row, column=13).value
        if currency == "$":
            currency = "USD"
        elif currency == "€":
            currency = "EUR"
        # service start state
        start_date = circuit_ws.cell(row=cell.row, column=11).value
        if start_date is None:
            c_notes.value = "In service start date was blank."
            continue
        # service term
        term_length = circuit_ws.cell(row=cell.row, column=12).value
        if term_length is None:
            c_notes.value = "Term length was blank"
            continue
        if term_length == '':
            c_notes.value = "Term length was blank"
            continue
        # contract end date
        contract_end = datetime.datetime.strptime(str(start_date), '%Y-%m-%d %H:%M:%S') + relativedelta(months=+int(term_length))
        # circuit name to be used as cost group name
        cost_group_name = circuit_ws.cell(row=cell.row, column=2).value
        # mrc - monthly recurring charge
        monthly_charge = circuit_ws.cell(row=cell.row, column=14).value
        # Get Price per Mbps
        price_per = circuit_ws.cell(row=cell.row, column=9).value
        # ignore blanks
        if circuit_id is None:
            c_notes.value = "Circuit ID was blank"
            continue
        if circuit_id.find("DCS Internal") > -1:
            c_notes.value = "DCS Internal"
            continue
        if circuit_id.find("cross-connect") > -1:
            c_notes.value = "Cross connect"
            continue
        if connect_type is None:
            c_notes.value = "Type was blank"
            continue
        if source_site_name is None:
            c_notes.value = "Site name was blank"
            continue
        if provider_name is None:
            c_notes.value = "Provider was blank"
            continue
        if monthly_charge is None:
            c_notes.value = "MRC was blank"
            continue
        if monthly_charge == '':
            c_notes.value = "MRC was blank"
            continue
        if monthly_charge == '0':
            c_notes.value = "MRC was 0"
            continue
        if price_per == '0':
            c_notes.value = "Price per Megabit was 0"
            continue
        if price_per == '':
            c_notes.value = "Price per Megabit was blank"
            continue
        # ignore provider name of snow
        if provider_name.find("SNOW") > -1:
            c_notes.value = "Provider was SNOW"
            continue
        # boolean used to check if interfaces were found
        interface_bool = False
        # list of interface ids
        interface_ids = {}
        # go through each interface and see if there is a match with the circuit id
        for row2 in interfaces_ws.iter_rows(min_col=5,max_col=5):
            for cell2 in row2:
                # ignore blank descriptions
                interface_desc = interfaces_ws.cell(row=cell2.row, column=5).value
                if interface_desc is None:
                    continue
                # see if the site name matches
                interface_site = interfaces_ws.cell(row=cell2.row, column=7).value
                # Remove white space in circuit id
                circuit_id = circuit_id.replace(" ", "")
                # check if circuit id is found
                if interface_desc.find(f"ref={circuit_id}") > -1:
                    interface_bool = True
                    if connect_type != "Internet":
                        if source_site_name.find(interface_site) > -1:
                            # define key variables
                            device_id = interfaces_ws.cell(row=cell2.row, column=2).value
                            interface_ids["aside_device"] = device_id
                            interface_id = interfaces_ws.cell(row=cell2.row, column=3).value
                            interface_ids["aside_interface"] = interface_id
                        else:
                            device_id = interfaces_ws.cell(row=cell2.row, column=2).value
                            interface_ids["bside_device"] = device_id
                            interface_id = interfaces_ws.cell(row=cell2.row, column=3).value
                            interface_ids["bside_interface"] = interface_id
                    else:
                        device_id = interfaces_ws.cell(row=cell2.row, column=2).value
                        interface_ids["internet_device"] = device_id
                        interface_id = interfaces_ws.cell(row=cell2.row, column=3).value
                        interface_ids["internet_interface"] = interface_id

        # was an interface found then add
        if interface_bool:
            # match cell headers to values
            c_connect_type = output_ws.cell(row = c_number, column = 1)
            c_provider = output_ws.cell(row = c_number, column = 2)
            c_cost_group_name = output_ws.cell(row = c_number, column = 3)
            c_cost_formula = output_ws.cell(row = c_number, column = 4)
            c_end_date = output_ws.cell(row = c_number, column = 5)
            c_bandwidth = output_ws.cell(row = c_number, column = 6)
            c_charge = output_ws.cell(row = c_number, column = 8)
            c_currency = output_ws.cell(row = c_number, column = 12)
            c_interface_id = output_ws.cell(row = c_number, column = 16)
            c_contract_id = output_ws.cell(row = c_number, column = 13)
            c_circuit_id = output_ws.cell(row = c_number, column = 14)
            c_billing_start_date = output_ws.cell(row = c_number, column = 15)
            c_device_id = output_ws.cell(row = c_number, column = 17)
            c_price_per_mb = output_ws.cell(row = c_number, column = 11)
            c_connect_type.value = circuit_type
            c_provider.value = provider_name
            c_cost_group_name.value = cost_group_name
            c_cost_formula.value = "Commit (Blended)"
            c_end_date.value = contract_end.strftime("%Y-%m-%d")
            c_bandwidth.value = bandwidth
            c_price_per_mb.value = price_per
            c_charge.value = monthly_charge
            c_currency.value = currency
            c_contract_id.value = circuit_ws.cell(row=cell.row, column=8).value
            c_billing_start_date.value = start_date.day
            c_circuit_id.value = circuit_id
            c_notes.value = "Circuit Added"
            c_number = c_number + 1
            if connect_type == "Internet":
                c_device_id.value = interface_ids["internet_device"]
                c_interface_id.value = interface_ids["internet_interface"]
            elif "aside_interface" in interface_ids:
                c_device_id.value = interface_ids["aside_device"]
                c_interface_id.value = interface_ids["aside_interface"]
            else:
                c_device_id.value = interface_ids["bside_device"]
                c_interface_id.value = interface_ids["bside_interface"]
        else:
            c_notes.value = "interface was not found"
            continue
output_workbook.save('./connectivit_costs_bulk_import_saved.xlsx')
circuit_workbook.save('./circuits.xlsx')
